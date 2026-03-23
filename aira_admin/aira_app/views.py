from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .models import ChatbotQuery, DetailImage, InventoryItem
from django.db.models import Q
from openpyxl import Workbook, load_workbook
from django.utils import timezone
from django.conf import settings
from datetime import datetime, date
from collections import Counter, defaultdict
from django.http import JsonResponse
from gtts import gTTS
import uuid
import re
import os
import json
from .utils import log_interaction_to_excel


ADMIN = "admin"
PASSWORD = "admin"

USER = "user"
PASS = "user"

def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        if username == ADMIN and password == PASSWORD:
            request.session["is_logged_in"] = True
            return redirect("home")
        elif username == USER and password == PASS:
            request.session["is_logged_in"] = True
            return redirect("userhome")
        else:
            return render(request, "login.html", {"error": "Invalid username or password"})
    return render(request, "login.html")


def logout_view(request):

    request.session.flush()  # This clears all session data

    return redirect("login")


def home_view(request):
    if not request.session.get("is_logged_in"):
        return redirect("login")
    return render(request, "home.html")


def query_list(request):
    if not request.session.get("is_logged_in"):
        return redirect("login")

    search_term = request.GET.get("search", "").strip()

    queries = ChatbotQuery.objects.all()
    if search_term:
        queries = queries.filter(query__icontains=search_term)

    return render(
        request,
        "query_list.html",
        {
            "queries": queries,
            "search": search_term,   # pass back to template
        },
    )



def add_query(request):
    if not request.session.get("is_logged_in"):
        return redirect("login")

    error_message = None

    if request.method == "POST":
        query = request.POST.get("query", "").strip()
        response = request.POST.get("response", "").strip()


        if ChatbotQuery.objects.filter(query__iexact=query).exists():
            error_message = "The query already exists."
        else:
            ChatbotQuery.objects.create(query=query, response=response)
            return redirect("query_list")

    return render(request, "add_query.html", {
        "error_message": error_message
    })

def edit_query(request, pk):
    if not request.session.get("is_logged_in"):
        return redirect("login")

    query_obj = get_object_or_404(ChatbotQuery, pk=pk)
    error_message = None

    if request.method == "POST":
        new_query = request.POST.get("query", "").strip()
        new_response = request.POST.get("response", "").strip()


        if ChatbotQuery.objects.filter(query__iexact=new_query).exclude(pk=pk).exists():
            error_message = "The query already exists."
        else:
            query_obj.query = new_query
            query_obj.response = new_response
            query_obj.save()
            return redirect("query_list")

    return render(request, "edit_query.html", {
        "query_obj": query_obj,
        "error_message": error_message
    })

def delete_query(request, pk):
    if not request.session.get("is_logged_in"):
        return redirect("login")
    query_obj = get_object_or_404(ChatbotQuery, pk=pk)
    if request.method == "POST":
        query_obj.delete()
        return redirect("query_list")
 
    query_obj.delete()
    return redirect("query_list")

# Path to Excel log file
LOG_FILE_PATH = os.path.join(settings.BASE_DIR, "aira_query_log.xlsx")


def log_interaction_to_excel(user_query: str, matched_query: str, answer: str):
    """
    Append one row to an Excel file with:
    [timestamp, user_query, matched_query, answer]
    """
    timestamp = timezone.now().strftime("%Y-%m-%d %H:%M:%S")

    # If file exists, open it; otherwise create a new workbook with header
    if os.path.exists(LOG_FILE_PATH):
        wb = load_workbook(LOG_FILE_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "AIRA Logs"
        # Header row
        ws.append(["Timestamp", "User Query", "Matched Query", "AIRA Answer"])

    # Append the new row
    ws.append([timestamp, user_query, matched_query or "", answer])

    # Save the workbook
    wb.save(LOG_FILE_PATH)
    
    

def load_logs_from_excel():
    """
    Read all rows from aira_query_log.xlsx and return a list of dicts:
    {
      'timestamp': datetime,
      'user_query': str,
      'matched_query': str,
      'answer': str,
      'matched': bool
    }
    """
    if not os.path.exists(LOG_FILE_PATH):
        return []

    wb = load_workbook(LOG_FILE_PATH)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) <= 1:
        return []

    header = rows[0]
    data_rows = rows[1:]

    logs = []
    for r in data_rows:
        # Expecting: [Timestamp, User Query, Matched Query, AIRA Answer]
        ts_val, user_q, matched_q, answer = r

        # Handle timestamp – may come as string or datetime
        if isinstance(ts_val, datetime):
            ts = ts_val
        else:
            try:
                ts = datetime.strptime(str(ts_val), "%Y-%m-%d %H:%M:%S")
            except Exception:
                # fallback: skip row if timestamp is invalid
                continue

        user_q = user_q or ""
        matched_q = matched_q or ""
        answer = answer or ""

        logs.append({
            "timestamp": ts,
            "user_query": user_q,
            "matched_query": matched_q,
            "answer": answer,
            "matched": bool(matched_q.strip()),
        })

    return logs


def analytics_view(request):
    if not request.session.get("is_logged_in"):
        return redirect("login")

    # ----- 1. Load all logs -----
    logs = load_logs_from_excel()

    # ----- 2. Read filter parameters (GET) -----
    start_date_str = request.GET.get("start_date", "")
    end_date_str = request.GET.get("end_date", "")
    status = request.GET.get("status", "all")        # all / matched / unmatched
    period = request.GET.get("period", "daily")      # daily / weekly / monthly

    # Parse date filters
    start_date = None
    end_date = None
    try:
        if start_date_str:
            start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
        if end_date_str:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
    except ValueError:
        start_date = None
        end_date = None

    # ----- 3. Apply filters to logs -----
    filtered_logs = []
    for log in logs:
        d = log["timestamp"].date()

        if start_date and d < start_date:
            continue
        if end_date and d > end_date:
            continue

        if status == "matched" and not log["matched"]:
            continue
        if status == "unmatched" and log["matched"]:
            continue

        filtered_logs.append(log)

    # ----- 4. Aggregate: most frequent queries -----
    # Use user_query text; you could also use matched_query instead.
    query_counter = Counter(l["user_query"] for l in filtered_logs if l["user_query"].strip())
    top_queries = query_counter.most_common(10)

    top_queries_labels = [q for q, _ in top_queries]
    top_queries_counts = [c for _, c in top_queries]

    # ----- 5. Aggregate: interaction trends over time -----
    time_buckets = defaultdict(int)

    for log in filtered_logs:
        ts = log["timestamp"]
        if period == "weekly":
            iso_year, iso_week, _ = ts.isocalendar()
            key = f"{iso_year}-W{iso_week:02d}"
        elif period == "monthly":
            key = f"{ts.year}-{ts.month:02d}"
        else:  # daily
            key = ts.date().isoformat()
        time_buckets[key] += 1

    # Sort time buckets by key (time)
    sorted_keys = sorted(time_buckets.keys())
    time_labels = sorted_keys
    time_counts = [time_buckets[k] for k in sorted_keys]

    # ----- 6. Summary stats -----
    total_queries = len(filtered_logs)
    matched_count = sum(1 for l in filtered_logs if l["matched"])
    unmatched_count = total_queries - matched_count

    filtered_logs_sorted = sorted(filtered_logs, key=lambda x: x["timestamp"], reverse=True)
    recent_logs = filtered_logs_sorted[:20]

    # ----- 7. Convert for Chart.js (JSON) -----
    context = {
        "start_date": start_date_str,
        "end_date": end_date_str,
        "status": status,
        "period": period,

        "total_queries": total_queries,
        "matched_count": matched_count,
        "unmatched_count": unmatched_count,

        "top_queries_labels_json": json.dumps(top_queries_labels),
        "top_queries_counts_json": json.dumps(top_queries_counts),
        "time_labels_json": json.dumps(time_labels),
        "time_counts_json": json.dumps(time_counts),

        "recent_logs": recent_logs,
    }

    return render(request, "analytics.html", context)



def generate_tts_audio(text: str) -> str:
    """
    Generate an MP3 file for the given text using gTTS
    and return its MEDIA URL.
    """
    if not text or not text.strip():
        return ""

    # Create directory in media folder
    tts_dir = os.path.join(settings.MEDIA_ROOT, "tts")
    if not os.path.exists(tts_dir):
        os.makedirs(tts_dir, exist_ok=True)

    filename = f"{uuid.uuid4().hex}.mp3"
    filepath = os.path.join(tts_dir, filename)

    try:
        tts = gTTS(text=text, lang="en", tld="com")
        tts.save(filepath)
        return settings.MEDIA_URL + "tts/" + filename
    except Exception as e:
        print(f"TTS Error: {e}")
        return ""



STOPWORDS = {
    "give", "get", "show", "tell", "please", "the", "a", "an", 
    "me", "us", "about", "details", "detail", "info", "information"
}

def normalize(text: str) -> str:
    text = text.lower()
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    return re.sub(r"\s+", " ", text).strip()



def userhome_view(request):
    # Optional: Authentication check
    if not request.session.get("is_logged_in"):
        return redirect("login")

    context = {
        "user_query": "",
        "answer": "",
        "suggestions": [],
        "tts_url": "",
    }

    if request.method == "POST":
        user_query_raw = request.POST.get("user_query", "").strip()
        context["user_query"] = user_query_raw

        if not user_query_raw:
            return render(request, "userhome.html", context)

        answer_text = ""
        found_queries = []

        # 1. Exact match
        exact = ChatbotQuery.objects.filter(query__iexact=user_query_raw).first()
        if exact:
            answer_text = exact.response
            found_queries = [exact]
        else:
            # 2. Partial match (Contains)
            contains = ChatbotQuery.objects.filter(query__icontains=user_query_raw)[:5]
            if contains.exists():
                answer_text = contains[0].response
                found_queries = list(contains)
            else:
                # 3. Keyword-based fuzzy match
                nq = normalize(user_query_raw)
                user_words = [w for w in nq.split() if w not in STOPWORDS and len(w) > 2]
                
                scored = []
                for q in ChatbotQuery.objects.all():
                    q_norm = normalize(q.query)
                    score = sum(1 for w in user_words if w in q_norm)
                    if score > 0:
                        scored.append((score, q))
                
                if scored:
                    # Sort by score descending
                    scored.sort(key=lambda x: x[0], reverse=True)
                    best_match = scored[0][1]
                    answer_text = best_match.response
                    # Extract top 3 suggestions
                    found_queries = [item[1] for item in scored[:3]]

        # Finalizing response
        if answer_text:
            context["answer"] = answer_text
            context["suggestions"] = found_queries
            context["tts_url"] = generate_tts_audio(answer_text)

            matched_query_text = found_queries[0].query if found_queries else ""
        else:
            answer_text = "I'm sorry, I couldn't find any information regarding that. Please try rephrasing your question."
            context["answer"] = answer_text
            matched_query_text = ""
            context["tts_url"] = generate_tts_audio(answer_text)

        
        log_interaction_to_excel(
            user_query=user_query_raw,
            matched_query=matched_query_text,
            answer=answer_text,
        )
    return render(request, "userhome.html", context)

def announcements_view(request):
    if not request.session.get("is_logged_in"):
        return redirect("login")

    announcement_images = DetailImage.objects.filter(
        category='ANNOUNCEMENT'
    ).order_by('-uploaded_at')

    return render(request, "announcements.html", {
        "announcement_images": announcement_images
    })



from .models import PlacementSummary

def placement_view(request):
    if not request.session.get("is_logged_in"):
        return redirect("login")

    placement_images = DetailImage.objects.filter(
        category='PLACEMENT'
    ).order_by('-uploaded_at')

    summary = PlacementSummary.objects.first()

    return render(request, "placement.html", {
        "placement_images": placement_images,
        "summary": summary,
    })



def add_details_view(request):
    if not request.session.get("is_logged_in"):
        return redirect("login")

    if request.method == "POST":
        category = request.POST.get("category")
        title = request.POST.get("title", "").strip()
        images = request.FILES.getlist("images")

        if category and images:
            for img in images:
                DetailImage.objects.create(
                    category=category,
                    title=title,
                    image=img
                )
            return redirect("add_details")  # go back to same page after upload

    # GET or after redirect: show existing images
    announcement_images = DetailImage.objects.filter(
        category='ANNOUNCEMENT'
    ).order_by('-uploaded_at')

    placement_images = DetailImage.objects.filter(
        category='PLACEMENT'
    ).order_by('-uploaded_at')

    return render(request, "add_details.html", {
        "announcement_images": announcement_images,
        "placement_images": placement_images,
    })


def delete_detail_image(request, pk):
    if not request.session.get("is_logged_in"):
        return redirect("login")

    img = get_object_or_404(DetailImage, pk=pk)

    if request.method == "POST":
        # Delete the file from storage first (optional but cleaner)
        if img.image and hasattr(img.image, 'path'):
            img.image.delete(save=False)

        img.delete()
        return redirect("add_details")

    # If someone hits the URL via GET, just redirect back
    return redirect("add_details")

def inventory_view(request):
    if not request.session.get("is_logged_in"):
        return redirect("login")

    search_term = request.GET.get("search", "").strip()

    items = InventoryItem.objects.all().order_by("name")

    if search_term:
        items = items.filter(
            Q(name__icontains=search_term) |
            Q(location__icontains=search_term) |
            Q(category__icontains=search_term) |   
            Q(notes__icontains=search_term)        
        )

    return render(request, "inventory.html", {
        "items": items,
        "search": search_term,   # pass back to template
    })

def add_inventory(request):
    if not request.session.get("is_logged_in"):
        return redirect("login")

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        category = request.POST.get("category", "OTHER")
        quantity = request.POST.get("quantity", "0")
        location = request.POST.get("location", "").strip()
        notes = request.POST.get("notes", "").strip()

        if name:
            try:
                qty = int(quantity)
            except ValueError:
                qty = 0
            InventoryItem.objects.create(
                name=name,
                category=category,
                quantity=qty,
                location=location,
                notes=notes,
            )
        return redirect("inventory_view")

    return render(request, "add_inventory.html")


def edit_inventory(request, pk):
    if not request.session.get("is_logged_in"):
        return redirect("login")

    item = get_object_or_404(InventoryItem, pk=pk)

    if request.method == "POST":
        item.name = request.POST.get("name", "").strip()
        item.category = request.POST.get("category", item.category)
        quantity = request.POST.get("quantity", str(item.quantity))
        item.location = request.POST.get("location", "").strip()
        item.notes = request.POST.get("notes", "").strip()

        try:
            item.quantity = int(quantity)
        except ValueError:
            pass

        item.save()
        return redirect("inventory_view")

    return render(request, "edit_inventory.html", {"item": item})


def delete_inventory(request, pk):
    if not request.session.get("is_logged_in"):
        return redirect("login")

    item = get_object_or_404(InventoryItem, pk=pk)
    item.delete()
    return redirect("inventory_view")

def inventory_details(request):
    if not request.session.get("is_logged_in"):
        return redirect("login")

    search_term = request.GET.get("search", "").strip()

    items = InventoryItem.objects.all().order_by("name")

    if search_term:
        items = items.filter(
            Q(name__icontains=search_term) |
            Q(location__icontains=search_term) |
            Q(category__icontains=search_term) |   # if category is a CharField with choices
            Q(notes__icontains=search_term)        # if you have notes field
        )

    return render(request, "inventory_details.html", {
        "items": items,
        "search": search_term,   # send current search back to template
    })
    
from .models import PlacementSummary

def edit_placement_summary(request):
    if not request.session.get("is_logged_in"):
        return redirect("login")

    summary, created = PlacementSummary.objects.get_or_create(
        id=1,
        defaults={
            "total_offers": 0,
            "highest_package": "₹0 LPA",
            "average_package": "₹0 LPA",
            "top_recruiters": "",
        }
    )

    if request.method == "POST":
        summary.total_offers = request.POST.get("total_offers")
        summary.highest_package = request.POST.get("highest_package")
        summary.average_package = request.POST.get("average_package")
        summary.top_recruiters = request.POST.get("top_recruiters")
        summary.save()

        return render(request, "edit_placement_summary.html", {
            "summary": summary,
            "updated": True
        })

    return render(request, "edit_placement_summary.html", {
        "summary": summary
    })
